# -*- coding: utf-8 -*-
"""生成 MelonAI Top3 选题信息表 → 桌面/结果/MelonAI_Top3选题信息表.xlsx"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DESKTOP = os.path.join(os.path.expanduser("~"), "Desktop")
OUT_DIR = os.path.join(DESKTOP, "结果")
os.makedirs(OUT_DIR, exist_ok=True)
OUT_PATH = os.path.join(OUT_DIR, "MelonAI_Top3选题信息表.xlsx")

HEADERS = [
    "排名", "视频标题（可发布）", "一句话选题", "新闻时间", "主题类别",
    "新闻简介", "核心内容要点", "官方已确认事实（✓）", "媒体报道·待核实（△）",
    "打工人视角（跟打工人有什么关系）", "带得走的行动", "时长建议",
    "完整视频脚本（口播文案）", "分镜脚本", "钩子方向",
    "官方来源URL（可直接访问）", "权威媒体来源URL", "信息真实性评级",
    "核实状态说明", "与账号内容联动", "制作备注",
]

def R(*cells):
    return cells

ROWS = []

# ============ Top 1: DeepSeek-V4-Flash ============
ROWS.append(R(
    "1",
    "《美国顶级AI刚降价80%，因为国产免费版追上来了》",
    "GPT-5.6 Luna 贵 60% 的活，国产免费模型就干得了。",
    "2026-07-31（正式版API上线公测）",
    "维度②该接触什么 · 大模型发布 / 国产开源",
    "DeepSeek 于 7 月 31 日发布 DeepSeek-V4-Flash 正式版（内部版本 V4-Flash-0731），API 上线公测，并上线国家超算互联网平台。模型为轻量化 MoE（总参 284B、激活 13B），结构与 4 月预览版一致、仅重新后训练，但 Agent 能力大幅增强，官方公布的 9 项 Agent 基准中多项超过三个月前的 V4-Pro 预览版。原生支持 OpenAI Responses API 格式、适配 Codex，无需改代码即可切换。价格极低，官方称百万 Token 长任务单 Token 算力仅为前代的 27%。",
    "1) 模型名 deepseek-v4-flash，现有集成不改代码自动升级\n"
    "2) 总参 284B / 激活 13B，轻量化 MoE，1M 超长上下文\n"
    "3) 支持思考模式/非思考模式切换\n"
    "4) 百万 Token 场景单 Token 计算量为前代 27%\n"
    "5) 原生支持 OpenAI Responses API 格式，针对性适配 Codex\n"
    "6) 官方 Agent 基准：Terminal Bench 2.1=82.7、NL2Repo=54.2、Cybergym=76.7、DeepSWE=54.4、Toolathlon verified=70.3、Agent Last Exam=25.2、Automation Bench(Public)=25.1、DSBench-FullStack=68.7、DSBench-Hard=59.6\n"
    "7) 本次仅升级 API 接口，V4-Pro API 及 APP/WEB 端未变；V4-Pro 正式版预计 2026 年 8 月初上线",
    "✓ 正式版 API 上线公测（官方更新日志，多权威媒体一致转述）\n"
    "✓ 模型参数 284B/13B、1M 上下文\n"
    "✓ 官方 9 项 Agent 基准分数\n"
    "✓ 兼容 OpenAI Responses API 格式、适配 Codex\n"
    "✓ V4-Pro 正式版 8 月初发布计划",
    "△ 公测定价（媒体转述官方文档，数字请以官网为准）：缓存命中输入约 $0.0028/百万 token、未命中约 $0.14/百万 token、输出约 $0.28/百万 token、最大输出 384K、并发 2500\n"
    "△ 'Artificial Analysis 智能指数 50 分（比 4 月高 10 分，比 Luna 低 1 分）'、'单任务成本比 GPT-5.6 Luna 低约 60%'——来自第三方评测媒体（东方网 AI 周报等），非官方口径\n"
    "△ 上线国家超算互联网平台（媒体报道）",
    "顶级干活能力 + 极低成本 → 普通打工人也能用上顶尖模型，选工具不用迷信国外付费产品；也直接拆掉'必须花钱买课/买会员才能用好 AI'的焦虑。",
    "今天就打开 chat.deepseek.com 或 DeepSeek 开放平台，用 V4-Flash 干一件真实工作上的活（写周报/整理数据/写代码），和现在用的工具对比一次。",
    "50 秒（对标《AI失控事件》新闻解读套路）",
    "【0-6s 钩子】美国顶级 AI 刚降价 80%，因为国产免费版，追到它脸上了。\n"
    "【6-30s 事件】7 月 31 号，DeepSeek 把 V4-Flash 正式版 API 开放公测了。它体量只有上一代的四分之一不到，可官方跑了九项智能体测试，好几项把三个月前的 V4-Pro 预览版都比下去了——干活的 AI，变小、变强、变快了。官方还直接兼容了 OpenAI 的接口，连写代码的 Codex 都能换过来用。价格呢？缓存命中输入每百万 token 只要几厘钱，输出不到 0.3 美元。\n"
    "【30-40s 打工人连接】翻译成人话：顶级的干活能力，现在普通人免费、或者几乎免费就能用。你之前觉得'贵的才好''国外才强'，这个认知该更新了。\n"
    "【40-50s 行动+收尾】今天下班，拿一件要交的活让 DeepSeek V4 干一遍，看它配不配。你觉得国产 AI 到什么水平了？评论区聊聊。",
    "0-6s | 大字'降价80%'+新闻标题拼接 | 钩子台词\n"
    "6-18s | 官方基准数据可视化/代码界面 | 事件前半（正式版/九项基准/超越Pro预览）\n"
    "18-30s | API 兼容 Codex 界面演示 | 事件后半（兼容 OpenAI 接口/价格）\n"
    "30-40s | 切打工人办公桌场景 | 连接（免费顶级能力，认知更新）\n"
    "40-50s | 出片提示字'今天让它干一件活' | 行动+CTA",
    "反常识钩子：从'美国顶级 AI 降价'切入，带出国产反超，制造'原来这么强'的认知翻转。",
    "官方更新日志：https://api-docs.deepseek.com/zh-cn/updates/\n"
    "官方更新日志（EN）：https://api-docs.deepseek.com/updates/\n"
    "官方产品：https://chat.deepseek.com ｜ 开放平台：https://platform.deepseek.com",
    "IT之家：https://www.ithome.com/0/984/116.htm\n"
    "腾讯新闻：https://news.qq.com/rain/a/20260731A0BE7D00\n"
    "凤凰科技：https://tech.ifeng.com/c/8vCwcaZT48q\n"
    "广州日报·花城+：https://huacheng.gz-cm.com/pages/2026/07/31/68b07c8c497547e8b4ee0ea6b33be259.html",
    "高（官方确认事实部分）／中（第三方评测数据需标注来源）",
    "官方来源 = DeepSeek API 官方文档更新日志（本环境无法直接抓取该域名，URL 来自搜索引擎索引，多权威媒体对同一组官方数字一致转述，可信度高）。'50 分/成本低 60%'等是第三方评测口径，视频里不要当官方数据说。",
    "与《AI失控事件》同属'新闻解读'套路，可形成 50s 新闻系列；直接命中话题地图'该接触什么'维度；成本选题可联动创意池《49元AI课割韭菜》。",
    "发布前在官网核对一次定价与基准分（以官方文档为准）；视频里说价格时用'约''以官网为准'，避免数据争议。",
))

# ============ Top 2: OpenAI Astra ============
ROWS.append(R(
    "2",
    "《AI花2000美元，连破10道数学世纪难题》",
    "上次 AI 自己跑出去闯祸，这次它自己干了四天正事：10 道人类卡了多年的数学题。",
    "2026-07-31（官方博客发布，部分媒体口径 8/1）",
    "维度③该学什么 · AI Agent 能力 / 前沿突破",
    "OpenAI 官方博客《Ten advances in mathematics and theoretical computer science》首次公开下一代模型 Astra（官方称'our next major model'）的内部版本，在十个长期未决的数学/理论计算机科学难题上做出新结果，横跨高维几何、编码理论、算术电路复杂度、群论、算子代数、量子复杂性、格密码学、极值组合等。每个论证由模型生成、人类协作整理、并用 Lean 4 形式化验证，共发布 249 页手稿合集与 AI 推理过程旁白。官方称全部 token 成本约 2000 美元（按 Sol API 价格）。",
    "1) Astra 为 OpenAI '下一款主要模型'（next major model），本次为内部版本研究成果，非正式产品发布\n"
    "2) 10 项新结果清单：高维球堆积新上界、二进制/球面码指数级改进、非 sofic 群存在性构造、Connes 刚性猜想反例、永久式算术电路下界、量子平行重复定理、最近格向量问题难度、Ehrhart 体积猜想、multicolor Ramsey 数（解决 Erdős 问题 183）、极值数猜想（Erdős 146/180）\n"
    "3) 每个论证用 Lean 4 形式化验证，证书公开在 GitHub openai/ten-proofs\n"
    "4) 249 页手稿 PDF + 每个问题的 AI 推理过程旁白\n"
    "5) 总 token 成本约 $2000（按 Sol API 价格）\n"
    "6) 官方明确：AI 生成的证明应如实署名 AI 系统，呼应 Leiden 宣言\n"
    "7) 5 月 OpenAI 还曾公布 Erdős 单位距离猜想反例",
    "✓ 10 项数学/理论计算机新结果（官方博客）\n"
    "✓ Astra 为 'next major model' 的内部版本（官方博客原话）\n"
    "✓ 总 token 成本约 $2000（官方博客，按 Sol API 价格）\n"
    "✓ Lean 4 形式化验证 + GitHub 公开证书\n"
    "✓ 249 页手稿与推理旁白公开",
    "△ 'Altman 在华盛顿向监管层演示 Astra'（The Information 报道，36氪/华尔街见闻转述）\n"
    "△ '可能命名 GPT-6 或 GPT-5.7''发布时间未定''或成首批受美国新 AI 监管框架审查的模型'（媒体推测，官方未确认）\n"
    "△ '菲尔兹奖级'是部分数学家/媒体评价，专家同行评审尚未完成",
    "数学难题本身离打工人远，但内核是：AI 从'答一句'升级到'独立干一件长活'——以后你可以把整件工作交给 AI，而不是一句一句喂。接续上一期'AI 自己干活没人盯会闯祸'的线，这次是'AI 干长活立了功'。",
    "把一个能拆成'交代清楚就能跑'的小任务（如让它自己调研+整理成报告），整件委托给 AI Agent 试一次，观察它能自主跑多久、做到什么程度。",
    "50-60 秒（事件讲足、连接压到一句，沿用 AI 失控事件复盘经验）",
    "【0-6s 钩子】AI 花 2000 美元，把人类十年没解开的数学题，一连解开了十道。\n"
    "【6-34s 事件】这是 OpenAI 官方发的：他们的下一代模型 Astra，把十个数学和理论计算机领域的难题，一口气做出了新结果——高维几何、群论、量子复杂性、格密码，全在里面。每一个论证，都用 Lean 证明语言逐行验证过，249 页手稿全部公开。注意，人类不是没试过——有的题卡了不止十年。而 OpenAI 说，全部算下来 token 成本大约 2000 美元。\n"
    "【34-46s 打工人连接】你可能觉得数学题跟我没关。关系在这：上一期我说过，AI 自己干活没人盯会闯祸；这次它证明了自己能独立干完一整件事。以后你交给 AI 的，不是一句话，是一整件任务。\n"
    "【46-56s 行动+收尾】现在就试试：把一个讲清楚就能放手的小任务，整件交给 AI，别一句一句喂。你信 AI 能干长活吗？评论区聊聊。",
    "0-6s | 黑屏大字'2000美元×10道题' | 钩子台词\n"
    "6-20s | 论文/证明页面滚动、Lean 代码画面 | 事件前半（官方发布/十大领域）\n"
    "20-34s | '卡了不止十年''249页''2000美元'数据卡 | 事件后半（验证+成本）\n"
    "34-46s | 接上一期画面闪回+打工人桌面 | 连接（AI 能独立干整件事）\n"
    "46-56s | 出片提示字'整件交给AI' | 行动+CTA",
    "猎奇钩子：'2000 美元 × 10 道世纪难题'天生吸流量；并延续《AI失控事件》形成'Agent 从闯祸到立功'系列感。",
    "官方博客：https://openai.com/index/ten-advances-in-mathematics/\n"
    "Lean 证明证书（GitHub）：https://github.com/openai/ten-proofs",
    "36氪：https://36kr.com/p/3921682068172419\n"
    "华尔街见闻：https://wallstreetcn.com/articles/3778497\n"
    "凤凰科技：https://tech.ifeng.com/c/8vEF0vvks8g\n"
    "网易：https://m.163.com/news/article/L3AQAVSS00097U7T.html",
    "高（官方博客确认事实）／中（媒体对发布时间、命名的推测部分）",
    "官方来源 = openai.com 官方博客（本环境无法直接抓取该域名，URL 来自搜索引擎索引；博客内容经 36氪、华尔街见闻、Simon Willison 等多方一致转述，事实部分可信）。媒体说的'命名 GPT-6/演示给监管层'是推测，视频里要区分'官方确认'与'据报道'。",
    "接续《AI失控事件》'Agent 从闯祸到立功'系列；联动概念笔记《智能体 Agent》；命中话题地图'AI Agent 认知'簇。",
    "视频里数学题只点名方向不背细节（避免说错领域）；'2000 美元'是官方口径但被指不含失败尝试与训练成本，建议表述为'官方称约 2000 美元'；'菲尔兹奖级'这类第三方评价不入文案。",
))

# ============ Top 3: Kimi K3 开源 + 融资 ============
ROWS.append(R(
    "3",
    "《全球最大开源AI模型免费给你用，背后的公司要上市了》",
    "顶级能力免费开源 + 350 亿美元估值资本大战：你用的 Kimi，正走向上市。",
    "发布 2026-07-17 ｜ 全链条开源 2026-07-27 ｜ F 轮融资完成 2026-07-29",
    "维度⑥怎么思考工作价值 · 开源模型 / 商业资本",
    "月之暗面（Moonshot AI）发布 Kimi K3——全球首个开源的三万亿级别大模型：2.8 万亿（2.78T）总参数、激活约 104B、MoE 896 专家每 Token 激活 16+2 共享、100 万 token 上下文、原生视觉（MoonViT-V2）。7/27 兑现承诺全链条开源（权重+47 页技术报告+MoonEP/FlashKDA/AgentENV 三套 infra），前端代码竞技场 WebDev Arena 以 1678 Elo 登顶、为首个登顶的开源模型。7/29 完成超 35 亿美元 F 轮（投后估值 350 亿美元，认购超 3 倍提前关闭），并提前开启投前估值 500 亿美元的 G 轮（Pre-IPO），7/29 完成股份制改造，市场传闻最快 6 个月内港股上市。",
    "1) 总参 2.8T（2.78T）/ 激活 104B，MoE 896 专家激活 16+2 共享，1,048,576 token 上下文，原生视觉编码器 MoonViT-V2（4 亿参数）\n"
    "2) 全球首个开源 3 万亿级别模型，参数规模超越 1.6T 的 DeepSeek-V4 Pro\n"
    "3) 7/27 全链条开源：权重（Hugging Face 上线 30 分钟 4000+ 赞、登顶趋势榜）+ 47 页技术报告《KIMI K3: OPEN FRONTIER INTELLIGENCE》（GitHub）+ MoonEP/FlashKDA/AgentENV 三套 infra（MIT）\n"
    "4) 架构创新：KDA 线性注意力（1M 上下文最高 6.3 倍解码加速）、Attention Residuals、Stable LatentMoE、MoonClip 二阶优化器（训练 FLOPs 减半）\n"
    "5) License：专门的 Kimi K3 License（MaaS 连续 12 个月总收入超 2000 万美元需另签商业协议等门槛）\n"
    "6) 评测：GPQA Diamond 93.5%、ProgramBench 77.8% 第一、WebDev Arena 1678 Elo 开源第一登顶、Kimi Code Bench 2.0 推理成本仅 Claude Fable 5 的 38%\n"
    "7) API 定价：输入 $3/M、输出 $15/M、缓存命中输入 $0.3/M（与 Claude Sonnet 5 一致，较 Fable 5 低 70%）\n"
    "8) Day0 适配：阿里云（真武M890）、华为昇腾、摩尔线程、海光、Nebius、Baseten、Fireworks、Cognition(Devin)；vLLM/SGLang 首日支持",
    "✓ 开源事实（官方 Hugging Face / GitHub / kimi.com，多源一致）\n"
    "✓ 模型参数、架构、License、API 定价（技术报告）\n"
    "✓ 评测分数与 Day0 生态适配（技术报告+媒体）",
    "△ F 轮超 35 亿美元、投后估值 350 亿美元（7/29）——上海证券报、每经、科创板日报、澎湃等权威媒体报道，但公司官方暂未公开回应\n"
    "△ G 轮投前估值 500 亿美元、最快下周关闭；股份制改造；最快 6 个月内港股上市——市场消息/媒体报道",
    "两层：① 3 万亿参数顶级能力免费开源 → AI 只会越来越便宜，49 块的 AI 课真没必要买（直接消化创意池高优选题）；② 你天天用的 Kimi 背后是几百亿美元资本大战——免费背后有人替你付费，看懂资本才不会被'焦虑营销'收割。",
    "今天打开 Kimi（kimi.com）用 K3 试一个你正犹豫要不要付费的功能，对比一下免费开源到底能不能打。",
    "60-90 秒（事件+商业双线，可做成 90s）",
    "【0-8s 钩子】全球参数最大的开源 AI 模型，免费给你用——发布它的公司，估值 350 亿美元，正准备上市。\n"
    "【8-40s 事件】月之暗面的 Kimi K3，7 月 27 号把完整权重开源了：2.8 万亿参数、100 万 token 上下文、自带视觉。这是全球第一个开源的三万亿级别模型，比之前最大的开源模型，参数几乎翻倍。前端写代码的评测榜，它直接登顶——是第一个把闭源旗舰都比下去的开源模型。开源当天，阿里、华为的服务器就全适配好了。\n"
    "【40-62s 商业】但更疯的是资本：前几天它完成 F 轮，融了超 35 亿美元，估值 350 亿——认购超过目标三倍，提前关闭。马上又开 G 轮，投前估值 500 亿，最快下周关。有媒体算过，从去年底到现在，估值半年涨了十倍多。公司还悄悄改成了股份公司，市场都在传它要去港股上市。\n"
    "【62-80s 打工人连接】对咱打工人，两层意思。第一，顶级能力免费开源，AI 只会越来越便宜，那些 49 块的 AI 课，你真没必要买。第二，你天天用的 Kimi，背后是场几百亿美元的资本大战——免费的背后，有的是人替你付钱。\n"
    "【80-90s 行动+收尾】今天打开 Kimi，用 K3 试一个你正犹豫要不要付费的功能，对比一下。你还在为 AI 付什么钱？评论区聊聊。",
    "0-8s | 大字'2.8万亿参数'弹入 | 钩子台词\n"
    "8-22s | Hugging Face 页面/参数卡特写 | 事件前半（参数/开源/登顶）\n"
    "22-40s | 技术报告封面+生态适配 logo 墙 | 事件后半（技术报告/阿里华为适配）\n"
    "40-62s | 融资数据滚动卡（35亿美元/350亿/500亿） | 商业（F轮/估值/G轮/上市传闻）\n"
    "62-80s | 打工人场景+对照卡'免费 vs 49元课' | 连接（别买课/看懂资本）\n"
    "80-90s | 出镜定格+CTA | 行动+CTA",
    "双钩子：'全球最大开源模型免费'（猎奇）+ '背后的公司要上市了'（商业故事），并顺带埋反割韭菜态度。",
    "官方开源（Hugging Face）：https://huggingface.co/moonshotai/Kimi-K3\n"
    "官方技术报告/模型卡（GitHub）：https://github.com/MoonshotAI/Kimi-K3\n"
    "官方产品入口：https://kimi.com",
    "上海证券报（融资）：https://paper.cnstock.com/html/2026-07/30/content_2250105.htm\n"
    "每经（融资）：https://www.mrjjxw.com/articles/2026-07-30/4526846.html\n"
    "科创板日报：http://www.dongshihui.net/finance/jiedu/2026-07-30/23538.html\n"
    "投资界（融资）：https://www.pedaily.cn/first/168641.shtml\n"
    "月之暗面发布Kimi K3（品玩）：https://www.pingwest.com/w/315601\n"
    "开源Day0适配（东方财富）：https://finance.eastmoney.com/a/202607283823453572.html",
    "高（开源事实部分）／中高（融资部分为权威媒体报道，官方暂未回应，需标注'据报道'）",
    "官方来源 = Hugging Face + GitHub + kimi.com（本环境无法直接抓取，URL 来自搜索引擎索引，多源一致）。融资部分为上证报/每经/科创板日报/澎湃等权威媒体一致报道，但公司官方暂未回应——视频里融资数字必须说'据媒体报道/市场消息'，不能当成官方公告。",
    "直接消化创意池高优先级《49元AI课一天卖10万：打工人怎么不被割》——开源免费事实可为反割韭菜提供论据；联动'职场双轨制'判断框架；命中维度⑥'怎么思考工作价值'。",
    "开源部分可放心讲（可溯源）；融资部分一律加'据媒体报道'；'50 万亿美元''股价蒸发数千亿'等不实或夸张说法勿用；技术细节（KDA/MoonViT）点到即可，别让文案变技术课。",
))

# ============ 写入 Excel ============
wb = Workbook()
ws = wb.active
ws.title = "MelonAI Top3 选题"

# 样式
header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
body_font = Font(name="微软雅黑", size=9)
thin = Side(style="thin", color="B0B0B0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap_top = Alignment(wrap_text=True, vertical="top", horizontal="left")
center = Alignment(wrap_text=True, vertical="center", horizontal="center")

# 表头
ws.append(HEADERS)
for c, _ in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=c)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

# 数据
rank_fill = {1: PatternFill("solid", fgColor="FDEBD0"), 2: PatternFill("solid", fgColor="D6EAF8"), 3: PatternFill("solid", fgColor="D5F5E3")}
for i, row in enumerate(ROWS, start=2):
    ws.append(list(row))
    rank = ws.cell(row=i, column=1).value
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=i, column=c)
        cell.font = body_font
        cell.alignment = wrap_top
        cell.border = border
    ws.cell(row=i, column=1).alignment = center
    ws.cell(row=i, column=1).fill = rank_fill.get(rank, PatternFill())

# 列宽
widths = {
    1: 5, 2: 30, 3: 28, 4: 22, 5: 22, 6: 48, 7: 60, 8: 46, 9: 46,
    10: 44, 11: 40, 12: 16, 13: 70, 14: 70, 15: 36, 16: 46, 17: 46,
    18: 22, 19: 60, 20: 44, 21: 44,
}
for idx, w in widths.items():
    ws.column_dimensions[get_column_letter(idx)].width = w

# 行高（数据行给足）
ws.row_dimensions[1].height = 30
for r in range(2, 2 + len(ROWS)):
    ws.row_dimensions[r].height = 400

# 冻结首行 + 自动筛选
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{1 + len(ROWS)}"

wb.save(OUT_PATH)
print("OK saved:", OUT_PATH)
